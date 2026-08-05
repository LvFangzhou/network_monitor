"""服务器资产、连接确认和受控端口变更 API。"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from app.config import settings
from app.database import get_db
from app.models import Device, ServerAsset, ServerIPAddress, ServerNIC, ServerNetworkConnection, ServerPortChange, User
from app.routers.auth import get_current_active_user
from app.schemas.server import (
    ConnectionDecision, ConnectionDiscovery, PortChangeApproval, PortChangeCreate,
    PortChangeExecute, ServerAssetCreate, ServerAssetUpdate, ServerAssetWithNetworkCreate,
    ServerIPCreate, ServerNICCreate,
)
from app.utils.server_connections import (
    MIN_CONNECTION_CONFIDENCE,
    normalize_mac,
    precheck_port_change,
    score_connection_evidence,
)

router = APIRouter()


def _admin(user: User):
    if not user.is_superuser and user.read_only:
        raise HTTPException(403, "只读账号不能修改服务器资产或网络配置")


def _change_approver(user: User):
    if not user.is_superuser:
        raise HTTPException(403, "只有超级管理员可以审批或执行服务器网络变更")


def _asset_dict(item: ServerAsset, detail: bool = False):
    data = {column.name: getattr(item, column.name) for column in item.__table__.columns}
    for key, value in list(data.items()):
        if isinstance(value, datetime): data[key] = value.isoformat()
    data["datacenter_name"] = item.datacenter_ref.name if item.datacenter_ref else None
    data["nic_count"] = len(item.nics or [])
    data["connection_count"] = sum(len(nic.connections or []) for nic in item.nics or [])
    if detail:
        data["components"] = [{**{c.name: getattr(row, c.name) for c in row.__table__.columns}, "last_discovered_at": row.last_discovered_at.isoformat() if row.last_discovered_at else None} for row in item.components]
        data["nics"] = []
        for nic in item.nics:
            row = {c.name: getattr(nic, c.name) for c in nic.__table__.columns}
            row["ip_addresses"] = [{c.name: getattr(ip, c.name) for c in ip.__table__.columns} for ip in nic.ip_addresses]
            row["connections"] = [_connection_dict(link) for link in nic.connections]
            data["nics"].append(row)
    return data


def _connection_dict(item: ServerNetworkConnection):
    return {
        "id": item.id, "server_id": item.server_id, "server_name": item.server.name if item.server else None,
        "nic_id": item.nic_id, "nic_name": item.nic.name if item.nic else None,
        "mac_address": item.nic.mac_address if item.nic else None, "switch_device_id": item.switch_device_id,
        "switch_name": item.switch_device.name if item.switch_device else None,
        "switch_ip": item.switch_device.ip_address if item.switch_device else None, "switch_port": item.switch_port,
        "state": item.state, "confidence": item.confidence, "confidence_level": item.confidence_level,
        "evidence": item.evidence or [], "conflict_reasons": item.conflict_reasons or [],
        "last_discovered_at": item.last_discovered_at.isoformat() if item.last_discovered_at else None,
        "confirmed_by": item.confirmed_by, "confirmed_at": item.confirmed_at.isoformat() if item.confirmed_at else None,
        "confirmation_note": item.confirmation_note,
    }


def _nic_dict(item: ServerNIC):
    return {
        **{column.name: getattr(item, column.name) for column in item.__table__.columns},
        "server_name": item.server.name if item.server else None,
        "management_ip": item.server.management_ip if item.server else None,
        "ip_addresses": [
            {column.name: getattr(ip, column.name) for column in ip.__table__.columns}
            for ip in item.ip_addresses or []
        ],
    }


@router.get("")
def list_servers(search: Optional[str] = None, datacenter_id: Optional[int] = None, status: Optional[str] = None,
                 skip: int = 0, limit: int = Query(50, le=500), db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_active_user)):
    query = db.query(ServerAsset).options(joinedload(ServerAsset.datacenter_ref), joinedload(ServerAsset.nics).joinedload(ServerNIC.connections))
    if search:
        token = f"%{search.strip()}%"
        query = query.filter(or_(ServerAsset.name.ilike(token), ServerAsset.management_ip.ilike(token), ServerAsset.serial_number.ilike(token), ServerAsset.asset_tag.ilike(token)))
    if datacenter_id: query = query.filter(ServerAsset.datacenter_id == datacenter_id)
    if status: query = query.filter(ServerAsset.status == status)
    total = query.count()
    return {"total": total, "items": [_asset_dict(row) for row in query.order_by(ServerAsset.id.desc()).offset(skip).limit(limit).all()]}


@router.post("")
def create_server(payload: ServerAssetCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user)
    item = ServerAsset(**payload.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return _asset_dict(item)


@router.post("/with-network")
def create_server_with_network(payload: ServerAssetWithNetworkCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    """在同一事务内录入服务器、网卡和 IP，避免留下半成品资产。"""
    _admin(current_user)
    data = payload.model_dump(exclude={"nics"})
    try:
        item = ServerAsset(**data)
        db.add(item)
        db.flush()
        for nic_payload in payload.nics:
            nic_data = nic_payload.model_dump(exclude={"ip_addresses"})
            nic_data["mac_address"] = normalize_mac(nic_data["mac_address"])
            nic = ServerNIC(server_id=item.id, **nic_data)
            db.add(nic)
            db.flush()
            for ip_payload in nic_payload.ip_addresses:
                db.add(ServerIPAddress(
                    nic_id=nic.id,
                    **ip_payload.model_dump(),
                    source="manual",
                    last_discovered_at=datetime.now(timezone.utc),
                ))
        db.commit()
        server_id = item.id
    except (ValueError, IntegrityError) as exc:
        db.rollback()
        detail = str(getattr(exc, "orig", exc))
        raise HTTPException(400, f"服务器、网卡或 IP 数据无效：{detail}") from exc
    return get_server(server_id, db, current_user)


@router.get("/nics")
def list_nics(search: Optional[str] = None, network_type: Optional[str] = None,
              skip: int = 0, limit: int = Query(500, le=1000), db: Session = Depends(get_db),
              current_user: User = Depends(get_current_active_user)):
    query = db.query(ServerNIC).options(joinedload(ServerNIC.server), joinedload(ServerNIC.ip_addresses))
    if search:
        token = f"%{search.strip()}%"
        query = query.join(ServerAsset).filter(or_(
            ServerAsset.name.ilike(token), ServerAsset.management_ip.ilike(token),
            ServerNIC.name.ilike(token), ServerNIC.mac_address.ilike(token),
        ))
    if network_type:
        query = query.filter(ServerNIC.network_type == network_type)
    total = query.count()
    rows = query.order_by(ServerNIC.server_id.desc(), ServerNIC.id).offset(skip).limit(limit).all()
    return {"total": total, "items": [_nic_dict(row) for row in rows]}


@router.get("/connections")
def list_connections(state: Optional[str] = None, skip: int = 0, limit: int = Query(500, le=1000),
                     db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    query = db.query(ServerNetworkConnection).options(
        joinedload(ServerNetworkConnection.server), joinedload(ServerNetworkConnection.nic),
        joinedload(ServerNetworkConnection.switch_device),
    )
    if state:
        query = query.filter(ServerNetworkConnection.state == state)
    total = query.count()
    rows = query.order_by(ServerNetworkConnection.last_discovered_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [_connection_dict(row) for row in rows]}


@router.get("/port-changes")
def list_port_changes(status: Optional[str] = None, skip: int = 0, limit: int = Query(500, le=1000),
                      db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    query = db.query(ServerPortChange).options(
        joinedload(ServerPortChange.connection).joinedload(ServerNetworkConnection.server),
        joinedload(ServerPortChange.connection).joinedload(ServerNetworkConnection.switch_device),
    )
    if status:
        query = query.filter(ServerPortChange.status == status)
    total = query.count()
    rows = query.order_by(ServerPortChange.requested_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": [_change_dict(row) for row in rows]}


@router.get("/{server_id}")
def get_server(server_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    item = db.query(ServerAsset).options(joinedload(ServerAsset.datacenter_ref), joinedload(ServerAsset.components), joinedload(ServerAsset.nics).joinedload(ServerNIC.ip_addresses), joinedload(ServerAsset.nics).joinedload(ServerNIC.connections).joinedload(ServerNetworkConnection.switch_device)).filter(ServerAsset.id == server_id).first()
    if not item: raise HTTPException(404, "服务器不存在")
    return _asset_dict(item, True)


@router.put("/{server_id}")
def update_server(server_id: int, payload: ServerAssetUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user); item = db.get(ServerAsset, server_id)
    if not item: raise HTTPException(404, "服务器不存在")
    for key, value in payload.model_dump(exclude_unset=True).items(): setattr(item, key, value)
    db.commit(); db.refresh(item); return _asset_dict(item)


@router.post("/{server_id}/nics")
def add_nic(server_id: int, payload: ServerNICCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user)
    if not db.get(ServerAsset, server_id): raise HTTPException(404, "服务器不存在")
    data = payload.model_dump(); data["mac_address"] = normalize_mac(data["mac_address"])
    nic = ServerNIC(server_id=server_id, **data); db.add(nic); db.commit(); db.refresh(nic)
    return {c.name: getattr(nic, c.name) for c in nic.__table__.columns}


@router.post("/nics/{nic_id}/ip-addresses")
def add_ip(nic_id: int, payload: ServerIPCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user)
    if not db.get(ServerNIC, nic_id): raise HTTPException(404, "网卡不存在")
    item = ServerIPAddress(nic_id=nic_id, **payload.model_dump(), source="manual", last_discovered_at=datetime.now(timezone.utc))
    db.add(item); db.commit(); db.refresh(item); return {c.name: getattr(item, c.name) for c in item.__table__.columns}


@router.post("/connections/discover")
def discover_connection(payload: ConnectionDiscovery, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user)
    nic = db.get(ServerNIC, payload.nic_id); switch = db.get(Device, payload.switch_device_id)
    if not nic or nic.server_id != payload.server_id or not switch: raise HTTPException(400, "服务器、网卡或交换机关系无效")
    evidence = [row.model_dump() for row in payload.evidence]
    score, level, conflicts = score_connection_evidence(evidence)
    item = db.query(ServerNetworkConnection).filter_by(nic_id=payload.nic_id, switch_device_id=payload.switch_device_id, switch_port=payload.switch_port).first()
    if not item:
        item = ServerNetworkConnection(server_id=payload.server_id, nic_id=payload.nic_id, switch_device_id=payload.switch_device_id, switch_port=payload.switch_port)
        db.add(item)
    item.evidence, item.confidence, item.confidence_level, item.conflict_reasons = evidence, score, level, conflicts
    item.last_discovered_at = datetime.now(timezone.utc)
    if item.state not in {"confirmed", "rejected"}: item.state = "candidate"
    db.commit(); db.refresh(item); return _connection_dict(item)


@router.post("/connections/{connection_id}/confirm")
def confirm_connection(connection_id: int, payload: ConnectionDecision, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user); item = db.get(ServerNetworkConnection, connection_id)
    if not item: raise HTTPException(404, "连接不存在")
    if item.confidence < MIN_CONNECTION_CONFIDENCE:
        raise HTTPException(409, f"可信度低于 {MIN_CONNECTION_CONFIDENCE} 分，需补充证据后才能确认")
    if item.conflict_reasons:
        raise HTTPException(409, "连接仍有冲突证据，必须补充证据或人工修正后才能确认")
    item.state, item.confirmed_by, item.confirmed_at, item.confirmation_note = "confirmed", current_user.username, datetime.now(timezone.utc), payload.note
    item.conflict_reasons = []
    db.commit(); db.refresh(item); return _connection_dict(item)


@router.post("/connections/{connection_id}/reject")
def reject_connection(connection_id: int, payload: ConnectionDecision, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user); item = db.get(ServerNetworkConnection, connection_id)
    if not item: raise HTTPException(404, "连接不存在")
    item.state, item.confirmed_by, item.confirmed_at, item.confirmation_note = "rejected", current_user.username, datetime.now(timezone.utc), payload.note
    db.commit(); return _connection_dict(item)


@router.post("/port-changes")
def create_change(payload: PortChangeCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user); link = db.get(ServerNetworkConnection, payload.connection_id)
    if not link: raise HTTPException(404, "连接不存在")
    check = precheck_port_change(link, payload.requested_config)
    # 第一阶段不接受前端或发现结果冒充交换机现状。厂商只读适配器完成前，
    # 工单只能保存为预检查失败，不能进入审批或执行阶段。
    existing = {}
    check["errors"].append("尚未从交换机实时读取并解析端口现有配置，禁止生成可审批差异")
    check["passed"] = False
    item = ServerPortChange(connection_id=link.id, status="precheck_failed", requested_config=payload.requested_config, existing_config=existing, config_diff=[], precheck_result=check, requested_by=current_user.username, audit_events=[{"action": "created", "actor": current_user.username, "at": datetime.now(timezone.utc).isoformat(), "reason": payload.reason}])
    db.add(item); db.commit(); db.refresh(item); return _change_dict(item)


def _change_dict(item: ServerPortChange):
    data = {c.name: getattr(item, c.name) for c in item.__table__.columns}
    for key, value in list(data.items()):
        if isinstance(value, datetime): data[key] = value.isoformat()
    data["execution_enabled"] = settings.SERVER_NETWORK_CHANGE_EXECUTION_ENABLED
    if item.connection:
        data.update({
            "server_name": item.connection.server.name if item.connection.server else None,
            "switch_name": item.connection.switch_device.name if item.connection.switch_device else None,
            "switch_ip": item.connection.switch_device.ip_address if item.connection.switch_device else None,
            "switch_port": item.connection.switch_port,
        })
    return data


@router.post("/port-changes/{change_id}/approve")
def approve_change(change_id: int, payload: PortChangeApproval, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _change_approver(current_user); item = db.get(ServerPortChange, change_id)
    if not item: raise HTTPException(404, "变更工单不存在")
    if item.requested_by == current_user.username: raise HTTPException(409, "申请人与审批人必须分离")
    if item.status != "precheck_passed": raise HTTPException(409, "只有预检查通过的工单可审批")
    item.status = "approved" if payload.approved else "rejected"; item.approved_by = current_user.username; item.approved_at = datetime.now(timezone.utc); item.approval_note = payload.note
    db.commit(); return _change_dict(item)


@router.post("/port-changes/{change_id}/execute")
def execute_change(change_id: int, payload: PortChangeExecute, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _change_approver(current_user); item = db.get(ServerPortChange, change_id)
    if not item: raise HTTPException(404, "变更工单不存在")
    if item.status != "approved": raise HTTPException(409, "工单尚未审批")
    if not payload.confirm: raise HTTPException(400, "必须二次确认执行")
    if not settings.SERVER_NETWORK_CHANGE_EXECUTION_ENABLED:
        raise HTTPException(409, "生产配置执行开关当前关闭；差异预览和审批已保留，不会下发设备")
    raise HTTPException(501, "该交换机厂商的端口配置适配器尚未验收，拒绝执行")


@router.post("/import")
async def import_servers(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    _admin(current_user); raw = await file.read()
    if file.filename and file.filename.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
            sheet = load_workbook(io.BytesIO(raw), read_only=True, data_only=True).active
            values = list(sheet.iter_rows(values_only=True))
            headers = [str(value or "").strip() for value in (values[0] if values else [])]
            rows = [dict(zip(headers, row)) for row in values[1:]]
        except Exception as exc:
            raise HTTPException(400, f"Excel 解析失败：{exc}") from exc
    else:
        text = raw.decode("utf-8-sig"); rows = csv.DictReader(io.StringIO(text))
    created, errors = 0, []
    for number, row in enumerate(rows, start=2):
        try:
            with db.begin_nested():
                def text_value(*keys: str) -> str:
                    value = next((row.get(key) for key in keys if row.get(key) is not None), "")
                    return str(value).strip()

                name = text_value("name", "服务器名称")
                if not name:
                    raise ValueError("服务器名称为空")
                item = ServerAsset(
                    name=name,
                    management_ip=text_value("management_ip", "管理IP") or None,
                    serial_number=text_value("serial_number", "序列号") or None,
                    vendor=text_value("vendor", "厂商") or None,
                    model=text_value("model", "型号") or None,
                    rack=text_value("rack", "机柜") or None,
                    status=text_value("status", "状态") or "in_stock",
                )
                db.add(item)
                db.flush()
            created += 1
        except Exception as exc:
            errors.append({"row": number, "error": str(exc)})
    db.commit(); return {"created": created, "failed": len(errors), "errors": errors}
